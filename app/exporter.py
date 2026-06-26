"""导出器：把转换结果写入用友BIP「凭证查询卡片主表」模板。"""

from __future__ import annotations

import shutil

import openpyxl

from .template_spec import DATA_SHEET, HEADER_ROWS, NUM_COLUMNS, TEMPLATE_PATH


def export_rows(rows: list[list[str]], output_path: str, template_path: str = TEMPLATE_PATH) -> int:
    """复制模板骨架（保留 9 行表头），从第 10 行开始写入数据行。返回写入行数。"""
    shutil.copyfile(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb[DATA_SHEET]
    # 清掉可能残留的数据行
    if ws.max_row > HEADER_ROWS:
        ws.delete_rows(HEADER_ROWS + 1, ws.max_row - HEADER_ROWS)
    for i, row in enumerate(rows):
        r = HEADER_ROWS + 1 + i
        values = (list(row) + [""] * NUM_COLUMNS)[:NUM_COLUMNS]
        for c, val in enumerate(values, start=1):
            if val != "":
                ws.cell(row=r, column=c, value=val)
    wb.save(output_path)
    wb.close()
    return len(rows)
