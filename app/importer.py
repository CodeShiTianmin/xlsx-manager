"""导入器：把 导入模版(.xls/.xlsx) 的分录与 思路过程.xlsx 的字典写入 SQLite。"""

from __future__ import annotations

import os
from typing import Any

import openpyxl

from .db import Database


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _read_sheet_rows(path: str, sheet: str | int | None = None) -> tuple[list[str], list[list[str]]]:
    """读取工作表，返回 (表头, 数据行)。兼容 .xls 与 .xlsx。"""
    ext = os.path.splitext(path)[1].lower()
    rows: list[list[str]] = []
    if ext == ".xls":
        import xlrd

        wb = xlrd.open_workbook(path)
        sh = wb.sheet_by_index(sheet if isinstance(sheet, int) else 0) if not isinstance(sheet, str) else wb.sheet_by_name(sheet)
        for r in range(sh.nrows):
            rows.append([_clean(c) for c in sh.row_values(r)])
    else:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet] if isinstance(sheet, str) else wb.worksheets[sheet or 0]
        for row in ws.iter_rows(values_only=True):
            rows.append([_clean(c) for c in row])
        wb.close()
    if not rows:
        return [], []
    return rows[0], rows[1:]


def import_entries(db: Database, path: str, append: bool = False) -> int:
    """导入「导入模版」分录。日期与凭证字号为空时向下沿用（合并单元格场景）。"""
    if not append:
        db.clear("import_entries")
    _, data = _read_sheet_rows(path)
    source = os.path.basename(path)
    last_date = ""
    last_voucher = ""
    seq = 0
    inserted = 0
    for row in data:
        row = (row + [""] * 6)[:6]
        date, voucher_no, summary, subject, debit, credit = row
        if not any([date, voucher_no, summary, subject, debit, credit]):
            continue
        if date:
            last_date = date
        if voucher_no:
            last_voucher = voucher_no
        seq += 1
        db.add_import_entry(
            {
                "seq": seq,
                "date": date or last_date,
                "voucher_no": voucher_no or last_voucher,
                "summary": summary,
                "subject": subject,
                "debit": _norm_amount(debit),
                "credit": _norm_amount(credit),
                "source_file": source,
            }
        )
        inserted += 1
    return inserted


def _norm_amount(value: str) -> str:
    value = (value or "").replace(",", "").strip()
    if not value:
        return ""
    try:
        f = float(value)
        return str(int(f)) if f.is_integer() else str(f)
    except ValueError:
        return value


def _find_col(header: list[str], *names: str) -> int:
    for i, h in enumerate(header):
        for n in names:
            if h == n:
                return i
    return -1


def import_dictionaries(db: Database, path: str) -> dict[str, int]:
    """从 思路过程.xlsx 导入全部字典表。返回每张表导入的条数。"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = set(wb.sheetnames)
    wb.close()
    result: dict[str, int] = {}

    if "用友科目" in sheets:
        result["dict_yonbip_subject"] = _import_yonbip_subject(db, path)
    if "亿企科目" in sheets:
        result["dict_yiqi_subject"] = _import_yiqi_subject(db, path)
    if "客户" in sheets:
        result["dict_customer"] = _import_simple(db, path, "客户", "dict_customer", ("客户编码",), ("客户名称",))
    if "供应商" in sheets:
        result["dict_supplier"] = _import_simple(db, path, "供应商", "dict_supplier", ("供应商编码",), ("供应商名称",))
    if "银行变二级" in sheets:
        result["dict_bank"] = _import_simple(db, path, "银行变二级", "dict_bank", ("账户编码",), ("账户名称",))
    if "账簿" in sheets:
        result["dict_ledger"] = _import_simple(db, path, "账簿", "dict_ledger", ("账簿编码",), ("账簿名称",))
    if "项目" in sheets:
        result["dict_project"] = _import_project(db, path)
    if "费用项目" in sheets:
        result["dict_cost_item"] = _import_cost_item(db, path)
    return result


def _import_simple(db: Database, path: str, sheet: str, table: str, code_names, name_names) -> int:
    header, data = _read_sheet_rows(path, sheet)
    ci = _find_col(header, *code_names)
    ni = _find_col(header, *name_names)
    db.clear(table)
    n = 0
    for row in data:
        code = row[ci] if 0 <= ci < len(row) else ""
        name = row[ni] if 0 <= ni < len(row) else ""
        if not code and not name:
            continue
        db.upsert_dict_row(table, {"code": code or name, "name": name or code})
        n += 1
    return n


def _import_project(db: Database, path: str) -> int:
    header, data = _read_sheet_rows(path, "项目")
    ci = _find_col(header, "项目编码")
    ni = _find_col(header, "项目名称")
    cat = _find_col(header, "所属项目类别")
    db.clear("dict_project")
    n = 0
    for row in data:
        code = row[ci] if 0 <= ci < len(row) else ""
        if not code:
            continue
        db.upsert_dict_row(
            "dict_project",
            {
                "code": code,
                "name": row[ni] if 0 <= ni < len(row) else "",
                "category": row[cat] if 0 <= cat < len(row) else "",
            },
        )
        n += 1
    return n


def _import_cost_item(db: Database, path: str) -> int:
    header, data = _read_sheet_rows(path, "费用项目")
    ci = _find_col(header, "编码")
    ni = _find_col(header, "名称")
    tc = _find_col(header, "费用项目类型编码")
    tn = _find_col(header, "费用项目类型")
    db.clear("dict_cost_item")
    n = 0
    for row in data:
        code = row[ci] if 0 <= ci < len(row) else ""
        if not code:
            continue
        db.upsert_dict_row(
            "dict_cost_item",
            {
                "code": code,
                "name": row[ni] if 0 <= ni < len(row) else "",
                "type_code": row[tc] if 0 <= tc < len(row) else "",
                "type_name": row[tn] if 0 <= tn < len(row) else "",
            },
        )
        n += 1
    return n


def _import_yonbip_subject(db: Database, path: str) -> int:
    header, data = _read_sheet_rows(path, "用友科目")
    idx = {
        "level": _find_col(header, "科目层级"),
        "status": _find_col(header, "启用状态"),
        "code": _find_col(header, "编码"),
        "name": _find_col(header, "名称"),
        "element": _find_col(header, "会计要素"),
        "direction": _find_col(header, "余额方向"),
        "aux": _find_col(header, "辅助核算"),
    }
    db.clear("dict_yonbip_subject")
    n = 0
    for row in data:
        def g(key: str) -> str:
            i = idx[key]
            return row[i] if 0 <= i < len(row) else ""

        code = g("code")
        if not code:
            continue
        db.upsert_dict_row(
            "dict_yonbip_subject",
            {k: g(k) for k in idx},
        )
        n += 1
    return n


def _import_yiqi_subject(db: Database, path: str) -> int:
    """亿企科目对照：前 4 列为亿企侧(科目编码/名称/辅助核算/余额方向)，
    第 5-8 列为用友侧(科目层级/编码/名称/辅助核算)。"""
    _, data = _read_sheet_rows(path, "亿企科目")
    db.clear("dict_yiqi_subject")
    n = 0
    for row in data:
        row = (row + [""] * 8)[:8]
        yq_code, yq_name, yq_aux, direction, yy_level, yy_code, yy_name, yy_aux = row
        if not yq_code and not yy_code:
            continue
        db.upsert_dict_row(
            "dict_yiqi_subject",
            {
                "yq_code": yq_code,
                "yq_name": yq_name,
                "yq_aux": yq_aux,
                "direction": direction,
                "yy_level": yy_level,
                "yy_code": yy_code,
                "yy_name": yy_name,
                "yy_aux": yy_aux,
            },
        )
        n += 1
    return n
